"""Formatted multi-sheet Excel reporting for catalog monitoring runs."""

from __future__ import annotations

import json
from collections.abc import Iterable, Sequence
from datetime import UTC, datetime
from decimal import Decimal
from pathlib import Path
from urllib.parse import urlsplit

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.models import ComparisonResult, InvalidRecord, Product, RunSummary

TIMESTAMP_FORMAT = "yyyy-mm-dd hh:mm:ss"
DATE_FORMAT = "yyyy-mm-dd"
CURRENCY_FORMAT = "#,##0.00"
PERCENTAGE_FORMAT = "0.00%"
HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
ZEBRA_FILL = PatternFill(fill_type="solid", fgColor="F5F9FC")
SUMMARY_LABEL_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
SUCCESS_FILL = PatternFill(fill_type="solid", fgColor="D9EAD3")
WARNING_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
ALERT_FILL = PatternFill(fill_type="solid", fgColor="FCE4D6")
PRICE_DECREASE_FILL = PatternFill(fill_type="solid", fgColor="D9EAD3")
PRICE_INCREASE_FILL = PatternFill(fill_type="solid", fgColor="F4CCCC")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9E2EC"),
    right=Side(style="thin", color="D9E2EC"),
    top=Side(style="thin", color="D9E2EC"),
    bottom=Side(style="thin", color="D9E2EC"),
)
SHEET_TAB_COLORS = {
    "All Products": "1F4E78",
    "New Products": "548235",
    "Price Changes": "C55A11",
    "Availability Changes": "7030A0",
    "Removed Products": "C00000",
    "Invalid Records": "BF9000",
    "Run Summary": "0F6B78",
}
PRODUCT_HEADERS = [
    "Product ID",
    "Title",
    "Category",
    "Price",
    "Currency",
    "Availability",
    "Rating",
    "Description",
    "Image URL",
    "Product URL",
    "Scraped At",
]


def generate_catalog_report(
    all_products: Iterable[Product],
    comparison_result: ComparisonResult,
    invalid_records: Iterable[InvalidRecord],
    run_summary: RunSummary,
    *,
    reports_dir: Path = Path("reports"),
    started_at: datetime | None = None,
    finished_at: datetime | None = None,
    status: str = "completed",
) -> Path:
    """Create a formatted catalog report workbook and return its saved path."""

    products = list(all_products)
    invalid = list(invalid_records)
    start = started_at or datetime.now(UTC)
    finish = finished_at or datetime.now(UTC)
    workbook = Workbook()
    workbook.remove(workbook.active)

    _write_products_sheet(workbook, "All Products", products)
    _write_products_sheet(workbook, "New Products", comparison_result.new_products)
    _write_price_changes_sheet(workbook, comparison_result)
    _write_availability_changes_sheet(workbook, comparison_result)
    _write_removed_products_sheet(workbook, comparison_result.removed_products)
    _write_invalid_records_sheet(workbook, invalid)
    _write_run_summary_sheet(workbook, comparison_result, run_summary, start, finish, status)

    reports_dir.mkdir(parents=True, exist_ok=True)
    output_path = _report_path(reports_dir, start)
    workbook.save(output_path)
    return output_path


def _write_products_sheet(workbook: Workbook, title: str, products: Iterable[Product]) -> None:
    """Write a product sheet with all standard product fields."""

    worksheet = _create_worksheet(workbook, title, PRODUCT_HEADERS)
    for product in products:
        row = [
            product.product_id,
            product.title,
            product.category,
            product.price,
            product.currency,
            product.availability,
            product.rating,
            product.description,
            str(product.image_url) if product.image_url is not None else None,
            str(product.product_url),
            _excel_datetime(product.scraped_at),
        ]
        _append_row(
            worksheet,
            row,
            currency_columns=(4,),
            hyperlink_columns={
                9: str(product.image_url) if product.image_url is not None else None,
                10: str(product.product_url),
            },
            timestamp_columns=(11,),
        )
    _finalize_worksheet(worksheet)


def _write_price_changes_sheet(workbook: Workbook, result: ComparisonResult) -> None:
    """Write price change rows with decimal and percentage formatting."""

    worksheet = _create_worksheet(
        workbook,
        "Price Changes",
        ["Product", "Old Price", "New Price", "Difference", "Difference %", "Currency", "URL"],
    )
    for change in result.price_changes:
        percentage = (
            change.percentage_difference / Decimal("100")
            if change.percentage_difference is not None
            else None
        )
        row = [
            change.product_title,
            change.old_price,
            change.new_price,
            change.absolute_difference,
            percentage,
            change.currency,
            str(change.product_url),
        ]
        _append_row(
            worksheet,
            row,
            currency_columns=(2, 3, 4),
            percentage_columns=(5,),
            hyperlink_columns={7: str(change.product_url)},
        )
        _highlight_price_change(worksheet, worksheet.max_row, change.percentage_difference)
    _finalize_worksheet(worksheet)


def _write_availability_changes_sheet(workbook: Workbook, result: ComparisonResult) -> None:
    """Write availability change rows with readable product URLs."""

    worksheet = _create_worksheet(
        workbook,
        "Availability Changes",
        ["Product", "Previous Status", "Current Status", "URL"],
    )
    for change in result.availability_changes:
        row = [
            change.product_title,
            change.previous_status,
            change.current_status,
            str(change.product_url),
        ]
        _append_row(worksheet, row, hyperlink_columns={4: str(change.product_url)})
        _highlight_availability_change(worksheet, worksheet.max_row, change.current_status)
    _finalize_worksheet(worksheet)


def _write_removed_products_sheet(workbook: Workbook, products: Iterable[Product]) -> None:
    """Write removed products using values from their last known snapshot."""

    worksheet = _create_worksheet(
        workbook,
        "Removed Products",
        [
            "Product ID",
            "Title",
            "Category",
            "Last Price",
            "Currency",
            "Last Availability",
            "URL",
            "Last Seen At",
        ],
    )
    for product in products:
        row = [
            product.product_id,
            product.title,
            product.category,
            product.price,
            product.currency,
            product.availability,
            str(product.product_url),
            _excel_datetime(product.scraped_at),
        ]
        _append_row(
            worksheet,
            row,
            currency_columns=(4,),
            hyperlink_columns={7: str(product.product_url)},
            timestamp_columns=(8,),
        )
    _finalize_worksheet(worksheet)


def _write_invalid_records_sheet(workbook: Workbook, records: Iterable[InvalidRecord]) -> None:
    """Write invalid raw records with readable and structured validation errors."""

    worksheet = _create_worksheet(
        workbook,
        "Invalid Records",
        [
            "Source Index",
            "Product ID",
            "Title",
            "Raw Price",
            "Product URL",
            "Error Reason",
            "Error Details",
        ],
    )
    for source_index, record in enumerate(records, start=1):
        raw = record.raw_product
        row = [
            source_index,
            raw.product_id,
            raw.title,
            raw.price,
            raw.product_url,
            record.reason,
            json.dumps(record.errors, ensure_ascii=False),
        ]
        _append_row(worksheet, row, hyperlink_columns={5: raw.product_url})
    _finalize_worksheet(worksheet)


def _write_run_summary_sheet(
    workbook: Workbook,
    result: ComparisonResult,
    summary: RunSummary,
    started_at: datetime,
    finished_at: datetime,
    status: str,
) -> None:
    """Write collection counts, comparison counts, timing, and final run status."""

    worksheet = _create_worksheet(workbook, "Run Summary", ["Metric", "Value"])
    duration = finished_at - started_at
    rows: list[tuple[str, object]] = [
        ("Run date", _excel_datetime(started_at).date()),
        ("Start time", _excel_datetime(started_at)),
        ("Finish time", _excel_datetime(finished_at)),
        ("Duration", str(duration)),
        ("Products collected", summary.total_scraped),
        ("Valid products", summary.valid_products),
        ("Invalid products", summary.invalid_records),
        ("New products", len(result.new_products)),
        ("Removed products", len(result.removed_products)),
        ("Price changes", len(result.price_changes)),
        ("Availability changes", len(result.availability_changes)),
        ("Unchanged products", len(result.unchanged_products)),
        ("Run status", status),
    ]
    for label, value in rows:
        _append_row(worksheet, [label, value])
    worksheet["B2"].number_format = DATE_FORMAT
    worksheet["B3"].number_format = TIMESTAMP_FORMAT
    worksheet["B4"].number_format = TIMESTAMP_FORMAT
    _style_run_summary(worksheet)
    _finalize_worksheet(worksheet)


def _create_worksheet(workbook: Workbook, title: str, headers: list[str]) -> Worksheet:
    """Create one consistently formatted worksheet with a header row."""

    worksheet = workbook.create_sheet(title)
    worksheet.sheet_properties.tabColor = SHEET_TAB_COLORS[title]
    worksheet.sheet_view.showGridLines = False
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
    worksheet.row_dimensions[1].height = 28
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    return worksheet


def _append_row(
    worksheet: Worksheet,
    values: Sequence[object],
    *,
    currency_columns: tuple[int, ...] = (),
    percentage_columns: tuple[int, ...] = (),
    hyperlink_columns: dict[int, str | None] | None = None,
    timestamp_columns: tuple[int, ...] = (),
) -> None:
    """Append one row and apply reusable cell-level formatting rules."""

    worksheet.append(values)
    row_index = worksheet.max_row
    for cell in worksheet[row_index]:
        cell.alignment = CELL_ALIGNMENT
        cell.border = THIN_BORDER
    for column in currency_columns:
        worksheet.cell(row_index, column).number_format = CURRENCY_FORMAT
    for column in percentage_columns:
        worksheet.cell(row_index, column).number_format = PERCENTAGE_FORMAT
    for column in timestamp_columns:
        worksheet.cell(row_index, column).number_format = TIMESTAMP_FORMAT
    for column, url in (hyperlink_columns or {}).items():
        _set_hyperlink(worksheet.cell(row_index, column), url)


def _set_hyperlink(cell: Cell, url: str | None) -> None:
    """Turn a URL cell into a readable clickable hyperlink when possible."""

    if url is None:
        return
    cell.value = _display_url(url)
    cell.hyperlink = url
    cell.style = "Hyperlink"


def _finalize_worksheet(worksheet: Worksheet) -> None:
    """Apply readable widths, borders, and alternating rows across a worksheet."""

    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        longest_value = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_letter].width = min(max(longest_value + 2, 12), 50)
        for cell in column_cells[1:]:
            cell.alignment = CELL_ALIGNMENT
            if cell.row % 2 == 0 and cell.fill.fill_type is None:
                cell.fill = ZEBRA_FILL


def _highlight_price_change(
    worksheet: Worksheet,
    row_index: int,
    percentage_difference: Decimal | None,
) -> None:
    """Color price-change rows green for decreases and red for increases."""

    fill = (
        PRICE_DECREASE_FILL
        if percentage_difference is not None and percentage_difference < 0
        else PRICE_INCREASE_FILL
    )
    for cell in worksheet[row_index]:
        cell.fill = fill
        cell.font = Font(bold=cell.column in (1, 5))


def _highlight_availability_change(
    worksheet: Worksheet,
    row_index: int,
    current_status: str,
) -> None:
    """Use green for available products and amber for constrained availability."""

    fill = SUCCESS_FILL if current_status == "in_stock" else WARNING_FILL
    for cell in worksheet[row_index]:
        cell.fill = fill
        cell.font = Font(bold=cell.column in (1, 3))


def _style_run_summary(worksheet: Worksheet) -> None:
    """Style the summary as a compact, skimmable run dashboard."""

    status_cell = worksheet["B14"]
    status_cell.fill = SUCCESS_FILL if status_cell.value == "completed" else ALERT_FILL
    status_cell.font = Font(bold=True, color="1F1F1F")
    for row_index in range(2, worksheet.max_row + 1):
        label_cell = worksheet.cell(row_index, 1)
        value_cell = worksheet.cell(row_index, 2)
        label_cell.fill = SUMMARY_LABEL_FILL
        label_cell.font = Font(bold=True, color="1F1F1F")
        if row_index in (6, 7, 8, 9, 10, 11, 12, 13):
            value_cell.font = Font(bold=True, size=12, color="1F1F1F")
        if row_index in (8, 10, 11):
            value_cell.fill = ALERT_FILL if value_cell.value else SUCCESS_FILL


def _report_path(reports_dir: Path, started_at: datetime) -> Path:
    """Return a report path, adding a time suffix rather than overwriting a report."""

    date_part = _excel_datetime(started_at).strftime("%Y-%m-%d")
    base_path = reports_dir / f"catalog_report_{date_part}.xlsx"
    if not base_path.exists():
        return base_path

    time_part = _excel_datetime(started_at).strftime("%H%M%S")
    suffixed_path = reports_dir / f"catalog_report_{date_part}_{time_part}.xlsx"
    sequence = 1
    while suffixed_path.exists():
        suffixed_path = reports_dir / f"catalog_report_{date_part}_{time_part}_{sequence}.xlsx"
        sequence += 1
    return suffixed_path


def _display_url(url: str) -> str:
    """Return a compact readable label for a clickable URL cell."""

    parsed = urlsplit(url)
    return f"{parsed.netloc}{parsed.path}" or url


def _excel_datetime(value: datetime) -> datetime:
    """Convert a datetime to a UTC-naive value accepted by Excel."""

    if value.tzinfo is None:
        return value
    return value.astimezone(UTC).replace(tzinfo=None)
