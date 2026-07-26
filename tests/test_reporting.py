"""Tests for formatted multi-sheet Excel reporting."""

from datetime import UTC, datetime, timedelta
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from app.models import (
    AvailabilityChange,
    ComparisonResult,
    InvalidRecord,
    PriceChange,
    Product,
    RawProduct,
    RunSummary,
)
from app.services.reporting import (
    CURRENCY_FORMAT,
    PERCENTAGE_FORMAT,
    PRICE_INCREASE_FILL,
    PRODUCT_HEADERS,
    SHEET_TAB_COLORS,
    SUCCESS_FILL,
    ZEBRA_FILL,
    generate_catalog_report,
)

STARTED_AT = datetime(2026, 7, 25, 8, 0, tzinfo=UTC)
FINISHED_AT = STARTED_AT + timedelta(minutes=2, seconds=30)


def make_product(**overrides: object) -> Product:
    """Build a product suitable for report output tests."""

    values: dict[str, object] = {
        "product_id": "sku-001",
        "title": "Fixture Product",
        "category": "Fixtures",
        "price": Decimal("19.99"),
        "currency": "USD",
        "availability": "in_stock",
        "rating": 4.5,
        "description": "A detailed fixture product description.",
        "image_url": "https://example.test/images/fixture.jpg",
        "product_url": "https://example.test/products/fixture",
        "scraped_at": STARTED_AT,
    }
    values.update(overrides)
    return Product.model_validate(values)


def make_comparison_result(product: Product) -> ComparisonResult:
    """Build comparison data covering each report change worksheet."""

    removed = make_product(product_id="removed-001", title="Removed Product")
    return ComparisonResult(
        new_products=[product],
        removed_products=[removed],
        price_changes=[
            PriceChange(
                product_id=product.product_id,
                product_title=product.title,
                old_price=Decimal("10.00"),
                new_price=product.price,
                absolute_difference=Decimal("9.99"),
                percentage_difference=Decimal("99.90"),
                currency=product.currency,
                product_url=product.product_url,
            )
        ],
        availability_changes=[
            AvailabilityChange(
                product_id=product.product_id,
                product_title=product.title,
                previous_status="out_of_stock",
                current_status="in_stock",
                product_url=product.product_url,
            )
        ],
        unchanged_products=[make_product(product_id="unchanged-001")],
    )


def test_generate_catalog_report_creates_formatted_workbook(tmp_path: Path) -> None:
    """The report contains required sheets, headers, formatting, and summary values."""

    product = make_product()
    invalid_record = InvalidRecord(
        raw_product=RawProduct(
            product_id="bad-001",
            title="Bad Product",
            price="not-a-price",
            product_url="https://example.test/products/bad",
        ),
        reason="product validation failed: price is invalid",
        errors=[{"field": "price", "message": "invalid price", "type": "value_error"}],
    )
    report_path = generate_catalog_report(
        [product],
        make_comparison_result(product),
        [invalid_record],
        RunSummary(total_scraped=2, valid_products=1, invalid_records=1),
        reports_dir=tmp_path,
        started_at=STARTED_AT,
        finished_at=FINISHED_AT,
        status="completed",
    )

    workbook = load_workbook(report_path)
    expected_sheets = [
        "All Products",
        "New Products",
        "Price Changes",
        "Availability Changes",
        "Removed Products",
        "Invalid Records",
        "Run Summary",
    ]
    assert workbook.sheetnames == expected_sheets
    assert workbook["All Products"].iter_rows is not None
    assert [cell.value for cell in workbook["All Products"][1]] == PRODUCT_HEADERS
    for sheet_name in expected_sheets:
        worksheet = workbook[sheet_name]
        assert worksheet.freeze_panes == "A2"
        assert worksheet.auto_filter.ref is not None

    all_products = workbook["All Products"]
    assert all_products.sheet_properties.tabColor.rgb == f"00{SHEET_TAB_COLORS['All Products']}"
    assert all_products.sheet_view.showGridLines is False
    assert all_products["D2"].number_format == CURRENCY_FORMAT
    assert all_products["A2"].fill.fgColor.rgb == f"00{ZEBRA_FILL.fgColor.rgb[-6:]}"
    assert all_products["J2"].hyperlink is not None
    assert all_products["J2"].hyperlink.target == "https://example.test/products/fixture"
    assert all_products["J2"].value == "example.test/products/fixture"

    price_changes = workbook["Price Changes"]
    assert price_changes["B2"].number_format == CURRENCY_FORMAT
    assert price_changes["E2"].number_format == PERCENTAGE_FORMAT
    assert price_changes["G2"].hyperlink is not None
    assert price_changes["A2"].fill.fgColor.rgb == f"00{PRICE_INCREASE_FILL.fgColor.rgb[-6:]}"

    summary_values = {
        row[0].value: row[1].value
        for row in workbook["Run Summary"].iter_rows(min_row=2, values_only=False)
    }
    assert summary_values["Products collected"] == 2
    assert summary_values["Valid products"] == 1
    assert summary_values["Invalid products"] == 1
    assert summary_values["New products"] == 1
    assert summary_values["Removed products"] == 1
    assert summary_values["Price changes"] == 1
    assert summary_values["Availability changes"] == 1
    assert summary_values["Unchanged products"] == 1
    assert summary_values["Run status"] == "completed"
    assert workbook["Run Summary"]["B14"].fill.fgColor.rgb == f"00{SUCCESS_FILL.fgColor.rgb[-6:]}"


def test_generate_catalog_report_handles_empty_data_and_avoids_overwriting(tmp_path: Path) -> None:
    """Empty sheets remain valid and repeated same-day reports receive a suffix."""

    result = ComparisonResult()
    summary = RunSummary()
    first_path = generate_catalog_report(
        [],
        result,
        [],
        summary,
        reports_dir=tmp_path,
        started_at=STARTED_AT,
        finished_at=FINISHED_AT,
        status="completed",
    )
    second_path = generate_catalog_report(
        [],
        result,
        [],
        summary,
        reports_dir=tmp_path,
        started_at=STARTED_AT,
        finished_at=FINISHED_AT,
        status="completed",
    )

    assert first_path.name == "catalog_report_2026-07-25.xlsx"
    assert second_path.name == "catalog_report_2026-07-25_080000.xlsx"
    workbook = load_workbook(second_path)
    assert workbook["All Products"].max_row == 1
    assert workbook["Price Changes"].max_row == 1
    assert workbook["Invalid Records"].max_row == 1


def test_reporting_module_is_importable() -> None:
    """The reporting layer is available for later implementation."""
    from app.services import reporting

    assert reporting.__doc__
