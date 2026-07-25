"""Fixture-driven end-to-end workflow coverage using SQLite and Excel output."""

from __future__ import annotations

import json
import logging
from collections.abc import Callable
from pathlib import Path
from typing import cast

import pytest
from openpyxl import load_workbook
from playwright.async_api import Page

from app.config import Settings
from app.db.database import create_database_engine, create_session_factory, initialize_database
from app.db.repositories import CatalogRepository
from app.models import RawProduct
from app.services.orchestration import CatalogRunService


class FixtureBrowser:
    """A no-op browser context manager used by the fixture-based integration test."""

    @property
    def page(self) -> Page:
        """Return an opaque page because the fixture scraper does not navigate it."""

        return cast(Page, object())

    async def __aenter__(self) -> FixtureBrowser:
        """Enter without launching Chromium."""

        return self

    async def __aexit__(
        self,
        exc_type: type[BaseException] | None,
        exc_value: BaseException | None,
        traceback: object | None,
    ) -> None:
        """Exit without browser resources."""

        del exc_type, exc_value, traceback


class FixtureScraper:
    """Return deterministic raw products from a local JSON fixture."""

    def __init__(self, products: list[RawProduct]) -> None:
        self._products = products

    async def collect_products(self) -> list[RawProduct]:
        """Return the current fixture catalog without network or browser access."""

        return self._products


def _load_fixture_products(name: str) -> list[RawProduct]:
    """Load one named deterministic catalog from the local JSON fixture."""

    fixture_path = Path(__file__).parent / "fixtures" / "end_to_end_catalogs.json"
    catalogs = json.loads(fixture_path.read_text(encoding="utf-8"))
    return [RawProduct.model_validate(item) for item in catalogs[name]]


def _scraper_factory(
    products: list[RawProduct],
) -> Callable[[Page, Settings, logging.Logger], FixtureScraper]:
    """Build an orchestration-compatible scraper factory for one fixture catalog."""

    def factory(_page: Page, _settings: Settings, _logger: logging.Logger) -> FixtureScraper:
        """Return a scraper that supplies the selected local fixture records."""

        return FixtureScraper(products)

    return factory


@pytest.mark.integration
async def test_two_run_workflow_persists_compares_and_reports(tmp_path: Path) -> None:
    """Two fixture catalogs produce expected snapshot changes and Excel report rows."""

    database_path = tmp_path / "catalog.db"
    settings = Settings(
        catalog_url="https://example.test/catalog",
        database_url=f"sqlite+aiosqlite:///{database_path}",
        output_dir=tmp_path / "reports",
        diagnostics_dir=tmp_path / "diagnostics",
    )
    engine = create_database_engine(settings.database_url)
    await initialize_database(engine)
    repository = CatalogRepository(create_session_factory(engine))
    logger = logging.getLogger("test.end_to_end")

    try:
        first_service = CatalogRunService(
            settings,
            repository,
            logger,
            browser_factory=FixtureBrowser,
            scraper_factory=_scraper_factory(_load_fixture_products("first")),
        )
        first_result = await first_service.run()

        second_service = CatalogRunService(
            settings,
            repository,
            logger,
            browser_factory=FixtureBrowser,
            scraper_factory=_scraper_factory(_load_fixture_products("second")),
        )
        second_result = await second_service.run()
    finally:
        await engine.dispose()

    assert first_result.status == "completed"
    assert first_result.report_path is not None
    assert first_result.report_path.exists()
    assert first_result.comparison_summary.new_products == 3

    assert second_result.status == "completed"
    assert second_result.report_path is not None
    assert second_result.comparison_summary.new_products == 1
    assert second_result.comparison_summary.removed_products == 1
    assert second_result.comparison_summary.price_changes == 1
    assert second_result.comparison_summary.availability_changes == 1

    workbook = load_workbook(second_result.report_path)
    assert workbook["New Products"].max_row == 2
    assert workbook["Removed Products"].max_row == 2
    assert workbook["Price Changes"].max_row == 2
    assert workbook["Availability Changes"].max_row == 2
    assert workbook["New Products"]["B2"].value == "Delta Product"
    assert workbook["Removed Products"]["B2"].value == "Beta Product"
    assert workbook["Price Changes"]["A2"].value == "Alpha Product"
    assert workbook["Availability Changes"]["A2"].value == "Alpha Product"
